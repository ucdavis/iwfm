#!/usr/bin/env python
# hdf2zxlsx_gw.py
# Convert IWFM Groundwater Zone Budget HDF5 file to Excel workbook using zone definitions
# Copyright (C) 2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

import sys
import os
import numpy as np
from collections import defaultdict

try:
    import h5py
except ImportError:
    print("Error: h5py module not found")
    print("Install with: pip install h5py")
    sys.exit(1)

try:
    from loguru import logger
except ImportError:
    print("Error: loguru module not found")
    print("Install with: pip install loguru")
    sys.exit(1)

try:
    from iwfm.debug import setup_debug_logger
except ImportError:
    print("Error: iwfm.debug module not found")
    print("Make sure iwfm package is installed")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Error: openpyxl module not found")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def read_zone_definition(zone_file):
    """
    Read zone definition file

    Returns
    -------
    zextent : int
        1 = zones defined for horizontal plane (all layers)
        0 = different zones for each layer
    zone_info : dict
        {zone_id: zone_name}
    element_zones : dict
        if zextent==1: {element: zone}
        if zextent==0: {(element, layer): zone}
    """
    zone_info = {}
    element_zones = {}
    zextent = None

    with open(zone_file, 'r') as f:
        lines = f.readlines()

    # Find ZEXTENT
    for i, line in enumerate(lines):
        line = line.strip()
        if line and not line.startswith('C'):
            try:
                zextent = int(line.split()[0])
                break
            except:
                continue

    if zextent is None:
        raise ValueError("Could not find ZEXTENT value in zone definition file")

    # Find zone names section
    reading_zones = False
    reading_elements = False

    for line in lines:
        line_strip = line.strip()

        # Skip comments and empty lines
        if not line_strip or line_strip.startswith('C'):
            # Check for section headers in comments
            if 'ZID' in line_strip and 'ZNAME' in line_strip:
                reading_zones = True
                reading_elements = False
                continue
            elif 'IE' in line_strip and 'ZONE' in line_strip:
                reading_zones = False
                reading_elements = True
                continue
            continue

        # Read zone definitions
        if reading_zones and not reading_elements:
            parts = line_strip.split(None, 1)
            if len(parts) >= 2:
                try:
                    zone_id = int(parts[0])
                    zone_name = parts[1]
                    zone_info[zone_id] = zone_name
                except:
                    pass

        # Read element-zone assignments
        elif reading_elements:
            parts = line_strip.split()
            if len(parts) >= 2:
                try:
                    element = int(parts[0])
                    if zextent == 1:
                        # Format: IE ZONE
                        zone = int(parts[1])
                        element_zones[element] = zone
                    else:
                        # Format: IE LAYER ZONE
                        if len(parts) >= 3:
                            layer = int(parts[1])
                            zone = int(parts[2])
                            element_zones[(element, layer)] = zone
                except:
                    pass

    return zextent, zone_info, element_zones


def hdf2zxlsx_gw(hdf_file, zone_file, output_file,
                 area_fact=0.000022957, area_units='AC',
                 vol_fact=0.000022957, vol_units='ACFT',
                 verbose=False, debug=False):
    """
    Convert IWFM Groundwater Zone Budget HDF5 file to Excel workbook

    Parameters
    ----------
    hdf_file : str
        Path to input HDF5 zone budget file
    zone_file : str
        Path to zone definition file
    output_file : str
        Path to output Excel file (.xlsx)
    area_fact : float
        Area conversion factor (default: sq ft to acres)
    area_units : str
        Area units for output
    vol_fact : float
        Volume conversion factor (default: cu ft to acre-ft)
    vol_units : str
        Volume units for output
    verbose : bool
        Print progress messages
    debug : bool, default=False
        Enable debug output (more detailed than verbose)
    """
    import iwfm

    # Configure loguru logger for debug mode
    if debug:
        setup_debug_logger()  # Auto-detects script name

    if not os.path.exists(hdf_file):
        print(f"Error: HDF5 file '{hdf_file}' not found")
        sys.exit(1)

    if not os.path.exists(zone_file):
        print(f"Error: Zone definition file '{zone_file}' not found")
        sys.exit(1)

    if debug:
        logger.debug(f"Reading zone definition: {zone_file}")

    # Read zone definitions
    zextent, zone_info, element_zones = read_zone_definition(zone_file)

    if debug:
        logger.debug(f"Output file: {output_file}")
        logger.debug(f"ZEXTENT: {zextent}")
        logger.debug(f"Zones defined: {len(zone_info)}")
        logger.debug(f"Element assignments: {len(element_zones)}")

    # Create Excel workbook
    wb = openpyxl.Workbook()

    # First sheet is empty (keep default Sheet)
    ws_empty = wb.active
    ws_empty.title = "Sheet1"

    # Read HDF5 file
    if debug:
        logger.debug(f"Reading HDF5 file: {hdf_file}")

    with h5py.File(hdf_file, 'r') as f:
        # Get metadata
        attrs = f['Attributes'].attrs

        n_elements = attrs.get('SystemData%NElements', attrs.get('nLocations', 1))
        n_timesteps = attrs['NTimeSteps']
        n_layers = attrs.get('SystemData%NLayers', attrs.get('NLayers', 1))

        # Get time step info
        start_date = attrs['TimeStep%BeginDateAndTime'].decode('utf-8')
        delta_t = attrs['TimeStep%DeltaT']
        time_unit = attrs['TimeStep%Unit'].decode('utf-8')

        # Generate datetime objects for Excel
        datetime_objs = iwfm.generate_datetime_objects(start_date, n_timesteps, delta_t, time_unit)

        # Get descriptor
        descriptor = attrs.get('Descriptor', b'IWFM Groundwater Zone Budget').decode('utf-8')

        # Build column headers from component names in Layer_1
        component_names_raw = list(f['Layer_1'].keys())

        # Filter out non-data items (like 'FaceFlows')
        component_names_raw = [c for c in component_names_raw if c not in ['FaceFlows', 'Storage', 'VerticalFlows']]

        # Extract unique base component names
        full_headers = []
        for comp_name in component_names_raw:
            if '_Inflow' in comp_name:
                base_name = comp_name.replace('_Inflow (+)', '').strip()
                if base_name not in full_headers:
                    full_headers.append(base_name)

        # Sort to match expected order
        priority_order = ['GW Storage', 'Streams', 'Tile Drains', 'Subsidence',
                         'Deep Percolation', 'Constrained General Head BC',
                         'Small Watershed Baseflow', 'Small Watershed Percolation',
                         'Diversion Recoverable Loss', 'Bypass Recoverable Loss',
                         'Pumping by Element', 'Pumping by Well', 'Root Water Uptake']

        # Sort headers by priority, then alphabetically
        sorted_headers = []
        for pheader in priority_order:
            if pheader in full_headers:
                sorted_headers.append(pheader)
                full_headers.remove(pheader)

        # Add remaining headers alphabetically
        sorted_headers.extend(sorted(full_headers))
        full_headers = sorted_headers

        if debug:
            logger.debug(f"Elements: {n_elements=}")
            logger.debug(f"Layers: {n_layers=}")
            logger.debug(f"Time steps: {n_timesteps=}")
            logger.debug(f"Start date: {start_date=}")
            logger.debug(f"Time Unit: {time_unit=}")
            logger.debug(f"Delta t: {delta_t=}")

        # Get element areas (if available)
        zone_areas = defaultdict(float)
        if 'SystemData%ElementAreas' in f['Attributes']:
            elem_areas = f['Attributes/SystemData%ElementAreas'][:] * area_fact
        elif 'Areas' in f['Attributes']:
            elem_areas = f['Attributes/Areas'][:] * area_fact
        else:
            elem_areas = None

        if elem_areas is not None:
            # Calculate zone areas
            for elem_idx in range(n_elements):
                element = elem_idx + 1  # Elements are 1-indexed
                if zextent == 1:
                    zone = element_zones.get(element, -99)
                    if zone != -99:
                        zone_areas[zone] += elem_areas[elem_idx]
                else:
                    # For layer-specific zones, divide area by number of layers
                    for layer in range(1, n_layers + 1):
                        zone = element_zones.get((element, layer), -99)
                        if zone != -99:
                            zone_areas[zone] += elem_areas[elem_idx] / n_layers

        # Get element-to-column mappings for each layer
        elem_col_maps = {}
        for layer_idx in range(1, n_layers + 1):
            map_name = f'Layer{layer_idx}_ElemDataColumns'
            if map_name in f['Attributes']:
                elem_col_maps[layer_idx] = f[f'Attributes/{map_name}'][:]
            else:
                if verbose:
                    print(f"  Warning: {map_name} not found in attributes")

        # Get component names mapping
        full_comp_names = [n.decode('utf-8').strip() for n in f['Attributes/FullDataNames'][:]]

        if debug:
            logger.debug(f"Component mappings loaded: {len(full_comp_names)} components")

        # Initialize zone data storage
        zone_data = defaultdict(lambda: defaultdict(lambda: {'in': np.zeros(n_timesteps),
                                                               'out': np.zeros(n_timesteps)}))

        # Process data by layer and component
        for layer_idx in range(1, n_layers + 1):
            layer_name = f'Layer_{layer_idx}'

            if layer_name not in f or layer_idx not in elem_col_maps:
                if verbose:
                    print(f"  Warning: {layer_name} or mapping not found")
                continue

            if debug:
                logger.debug(f"Processing {layer_name}...")

            layer_group = f[layer_name]
            elem_col_map = elem_col_maps[layer_idx]

            # Process each component
            for comp_idx, comp_full_name in enumerate(full_comp_names):
                # Determine if this is inflow or outflow
                is_inflow = '_Inflow' in comp_full_name or '(+)' in comp_full_name
                flow_dir = 'in' if is_inflow else 'out'

                # Get the component base name
                comp_base = comp_full_name.replace('_Inflow (+)', '').replace('_Outflow (-)', '').strip()

                # Find matching header column index
                col_idx = None
                for i, header in enumerate(full_headers):
                    if comp_base == header:
                        col_idx = i
                        break

                if col_idx is None:
                    continue

                # Get data array
                dataset_path = f'{layer_name}/{comp_full_name}'
                if dataset_path not in f:
                    continue

                data_array = f[dataset_path][:]

                # Apply unit conversion
                data_array = data_array * vol_fact

                # Aggregate to zones using the element-to-column mapping
                for elem_idx in range(n_elements):
                    element = elem_idx + 1  # Elements are 1-indexed

                    # Get the column index for this element in this component's dataset
                    data_col = elem_col_map[comp_idx, elem_idx]

                    if data_col == 0:
                        # This element doesn't have data for this component
                        continue

                    # Convert from 1-based to 0-based index
                    data_col_idx = data_col - 1

                    # Determine which zone this element belongs to
                    if zextent == 1:
                        # Same zone for all layers
                        zone = element_zones.get(element, -99)
                    else:
                        # Different zones per layer
                        zone = element_zones.get((element, layer_idx), -99)

                    if zone == -99:
                        continue  # Skip elements not in any zone

                    # Add this element's data to the zone total
                    zone_data[zone][col_idx][flow_dir] += data_array[:, data_col_idx]

        if debug:
            logger.debug(f"Aggregation complete for {len(zone_data)} zones")

    # Determine unit labels
    if vol_units.upper() in ['ACFT', 'AC-FT', 'ACRE-FT', 'ACRE-FEET']:
        vol_label = 'ac.ft.'
    elif vol_units.upper() in ['AF']:
        vol_label = 'af'
    else:
        vol_label = vol_units.lower()

    if area_units.upper() in ['AC', 'ACRES']:
        area_label = 'acres'
    else:
        area_label = area_units.lower()

    # Write Excel sheets
    if debug:
        logger.debug(f"Writing Excel workbook: {output_file}")

    for zone_id in sorted(zone_data.keys()):
        zone_name = zone_info.get(zone_id, f'Zone{zone_id}')
        zone_area = zone_areas.get(zone_id, 0.0)

        if debug:
            logger.debug(f"Zone {zone_id}: {zone_name}")

        # Create new sheet for this zone (max 31 chars for sheet name)
        sheet_name = f"Zone{zone_id}_{zone_name}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Write title rows
        ws['A1'] = descriptor
        ws['A1'].font = Font(bold=True)

        budget_title = f"GROUNDWATER ZONE BUDGET IN {vol_label} FOR ZONE {zone_id} ({zone_name})"
        ws['A2'] = budget_title
        ws['A2'].font = Font(bold=True)

        area_title = f"ZONE AREA: {zone_area:,.2f} {area_label}"
        ws['A3'] = area_title

        # Build column headers with IN/OUT for each component
        # Row 5: component names (merged cells for each component)
        # We'll write headers starting at row 5
        col = 1
        ws.cell(row=5, column=col, value='Time')
        col += 1

        for header in full_headers:
            # Write component name spanning 2 columns
            ws.cell(row=5, column=col, value=header)
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)

            # Write IN/OUT sub-headers
            ws.cell(row=6, column=col, value='IN (+)')
            ws.cell(row=6, column=col+1, value='OUT (-)')

            col += 2

        # Add Discrepancy and Absolute Storage columns
        ws.cell(row=5, column=col, value='Discrepancy')
        ws.cell(row=6, column=col, value='(=)')
        col += 1

        ws.cell(row=5, column=col, value='Absolute Storage')
        ws.cell(row=6, column=col, value='')

        # Make header rows bold, centered, and wrap text
        for cell in ws[5]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[6]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data rows starting at row 7
        for time_idx in range(n_timesteps):
            row_num = time_idx + 7
            col = 1

            # Write datetime
            ws.cell(row=row_num, column=col, value=datetime_objs[time_idx])
            col += 1

            total_in = 0.0
            total_out = 0.0

            # Write IN/OUT values for each component
            for comp_idx in range(len(full_headers)):
                in_val = zone_data[zone_id][comp_idx]['in'][time_idx]
                out_val = zone_data[zone_id][comp_idx]['out'][time_idx]

                ws.cell(row=row_num, column=col, value=in_val)
                ws.cell(row=row_num, column=col+1, value=out_val)
                col += 2

                total_in += in_val
                total_out += out_val

            # Calculate discrepancy
            discrepancy = total_in - total_out

            # Get absolute storage from GW Storage component
            abs_storage = 0.0
            for comp_idx, header in enumerate(full_headers):
                if 'Storage' in header or 'STORAGE' in header:
                    abs_storage = zone_data[zone_id][comp_idx]['in'][time_idx] - zone_data[zone_id][comp_idx]['out'][time_idx]
                    break

            ws.cell(row=row_num, column=col, value=discrepancy)
            col += 1
            ws.cell(row=row_num, column=col, value=abs_storage)

        # Format column A (dates): mm/dd/yyyy format, width 11.0
        ws.column_dimensions['A'].width = 11.0
        for row_num in range(7, 7 + n_timesteps):
            ws.cell(row=row_num, column=1).number_format = 'mm/dd/yyyy'

        # Format remaining columns: number format with comma separator and 2 decimals, width 12.0
        total_cols = 1 + (len(full_headers) * 2) + 2  # Time + (IN/OUT pairs) + Disc + Storage
        for col_idx in range(2, total_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12.0
            for row_num in range(7, 7 + n_timesteps):
                ws.cell(row=row_num, column=col_idx).number_format = '#,##0.00'

    # Save workbook
    wb.save(output_file)

    if verbose:
        print(f"  Excel workbook written to: {output_file}")


if __name__ == '__main__':
    import argparse
    import iwfm.debug as idb

    parser = argparse.ArgumentParser(
        description='Convert IWFM Groundwater Zone Budget HDF5 to Excel workbook',
        formatter_class=argparse.RawDescriptionHelpFormatter)

    parser.add_argument('hdf_file', help='Input HDF5 zone budget file')
    parser.add_argument('zone_file', help='Zone definition file (.dat)')
    parser.add_argument('output_file', help='Output Excel file (.xlsx)')

    parser.add_argument('--area-fact', type=float, default=0.000022957,
                        help='Area conversion factor (default: 0.000022957 for sq ft to acres)')
    parser.add_argument('--area-units', type=str, default='AC',
                        help='Area units for output (default: AC)')

    parser.add_argument('--vol-fact', type=float, default=0.000022957,
                        help='Volume conversion factor (default: 0.000022957 for cu ft to acre-ft)')
    parser.add_argument('--vol-units', type=str, default='ACFT',
                        help='Volume units for output (default: ACFT)')

    parser.add_argument('--quiet', action='store_true', help='Suppress progress messages')

    parser.add_argument('--debug', action='store_true', help='Enable debug output')

    args = parser.parse_args()

    # Get file names
    if args.hdf_file:
        hdf_file = args.hdf_file
        if args.zone_file:
            zone_file = args.zone_file
        else:
            print("Error: ZBudget Zones file name needed")
            sys.exit(1)
        if args.output_file:
            output_file = args.output_file
        else:
            # Default: replace .hdf with .xlsx
            base_name = os.path.splitext(hdf_file)[0]
            output_file = f"{base_name}.xlsx"
    else:
        # Interactive mode
        print("IWFM Unsaturated Zone Budget HDF5 to Excel Converter")
        hdf_file = input('  Enter Groundwater ZBudget HDF5 file name: ')
        zone_file = input('  Enter ZBudget Zones file name: ')
        output_file = input('  Enter Excel output file name: ')


    idb.exe_time()  # initialize timer

    hdf2zxlsx_gw(args.hdf_file, args.zone_file, args.output_file,
                 area_fact=args.area_fact, area_units=args.area_units,
                 vol_fact=args.vol_fact, vol_units=args.vol_units,
                 verbose=not args.quiet, debug=args.debug)

    idb.exe_time()  # print execution time
