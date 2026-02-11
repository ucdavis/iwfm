# _openpyxl_backend.py
# openpyxl backend for Excel operations (cross-platform)
# Copyright (C) 2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

"""
openpyxl backend for Excel file operations.

This is the primary backend for iwfm.xls, providing cross-platform
Excel read/write support.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from iwfm.debug.logger_setup import logger


def create_workbook(filename=None):
    """Create a new Excel workbook.

    Parameters
    ----------
    filename : str, optional
        Path to save workbook to when save_workbook() is called.

    Returns
    -------
    workbook : openpyxl.Workbook
        New Excel workbook object.
    """
    logger.debug(f"Creating new workbook{f' for {filename}' if filename else ''}")
    wb = Workbook()
    wb._iwfm_filename = filename
    logger.info(f"Created new workbook{f': {filename}' if filename else ''}")
    return wb


def open_workbook(filename, data_only=False):
    """Open an existing Excel workbook.

    Parameters
    ----------
    filename : str
        Path to existing Excel file.
    data_only : bool, default=False
        If True, read cell values only (not formulas).

    Returns
    -------
    workbook : openpyxl.Workbook
        Opened Excel workbook object.

    Raises
    ------
    FileNotFoundError
        If the file does not exist.
    """
    import os
    if not os.path.exists(filename):
        raise FileNotFoundError(f"Excel file not found: {filename}")

    logger.debug(f"Opening workbook: {filename} (data_only={data_only})")
    wb = load_workbook(filename, data_only=data_only)
    wb._iwfm_filename = filename
    logger.info(f"Opened workbook: {filename} ({len(wb.sheetnames)} sheets)")
    logger.debug(f"Sheet names: {wb.sheetnames}")
    return wb


def save_workbook(workbook, filename=None):
    """Save workbook to file.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook to save.
    filename : str, optional
        Path to save to. If not provided, uses filename from
        create_workbook() or open_workbook().

    Raises
    ------
    ValueError
        If no filename provided and workbook has no associated filename.
    """
    save_path = filename or getattr(workbook, '_iwfm_filename', None)
    if not save_path:
        raise ValueError("No filename specified and workbook has no associated filename")

    logger.debug(f"Saving workbook to: {save_path}")
    workbook.save(save_path)
    logger.info(f"Saved workbook: {save_path}")


def close_workbook(workbook):
    """Close workbook and release resources.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook to close.
    """
    logger.debug("Closing workbook")
    workbook.close()
    logger.debug("Workbook closed")


def add_worksheet(workbook, name=None, position=None):
    """Add a new worksheet to workbook.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook to add sheet to.
    name : str, optional
        Name for the new sheet.
    position : int, optional
        Position to insert sheet (0-indexed). If None, adds at end.

    Returns
    -------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The newly created worksheet.
    """
    if position is None:
        ws = workbook.create_sheet(title=name)
    else:
        ws = workbook.create_sheet(title=name, index=position)
    logger.debug(f"Added worksheet: '{ws.title}'")
    return ws


def get_worksheet(workbook, name_or_index):
    """Get worksheet by name or index.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook to get sheet from.
    name_or_index : str or int
        Sheet name or 0-based index.

    Returns
    -------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The requested worksheet.
    """
    if isinstance(name_or_index, int):
        ws = workbook.worksheets[name_or_index]
    else:
        ws = workbook[name_or_index]
    logger.debug(f"Selected worksheet: '{ws.title}'")
    return ws


def write_cells(worksheet, data, start_row=1, start_col=1):
    """Write 2D array to worksheet.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to write to.
    data : list of lists
        2D array of data to write.
    start_row : int, default=1
        Starting row (1-indexed).
    start_col : int, default=1
        Starting column (1-indexed).
    """
    if not data:
        logger.debug("No data to write")
        return

    rows = len(data)
    cols = len(data[0]) if data else 0

    for i, row in enumerate(data):
        for j, value in enumerate(row):
            worksheet.cell(row=start_row + i, column=start_col + j, value=value)

    logger.debug(f"Wrote {rows}x{cols} cells to '{worksheet.title}' at ({start_row}, {start_col})")


def write_budget_data(workbook, budget_data):
    """Write IWFM budget data to workbook with standard formatting.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook to write to.
    budget_data : list
        List containing four lists:
        - loc_names : list of str - Location names
        - column_headers : list of lists - Column headers for each location
        - loc_values : list of DataFrames - Values for each location
        - titles : list of lists - Titles (3 items) for each location
    """
    loc_names, column_headers, loc_values, titles = budget_data

    logger.info(f"Writing budget data for {len(loc_names)} locations")

    for loc in range(len(loc_names)):
        # Create worksheet for this location
        ws = workbook.create_sheet(title=loc_names[loc])
        logger.debug(f"Creating budget sheet: '{loc_names[loc]}'")

        # Write titles (rows 1-4)
        for i, title in enumerate(titles[loc]):
            cell = ws.cell(row=i + 1, column=1, value=title)
            cell.font = Font(bold=True)

        # Write column headers (row 5)
        for i, header in enumerate(column_headers[loc]):
            cell = ws.cell(row=5, column=i + 1, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data (rows 6+)
        df = loc_values[loc]
        for i in range(df.shape[0]):
            # Date column (first column of dataframe)
            date_val = df.iloc[i, 0]
            if hasattr(date_val, 'strftime'):
                ws.cell(row=i + 6, column=1, value=date_val.strftime('%m/%d/%Y'))
            else:
                ws.cell(row=i + 6, column=1, value=date_val)
            ws.cell(row=i + 6, column=1).number_format = 'MM/DD/YYYY'

            # Data columns (remaining columns)
            for j in range(1, df.shape[1]):
                cell = ws.cell(row=i + 6, column=j + 1, value=df.iloc[i, j])
                cell.number_format = '#,##0.00'

        # Set column widths
        for col in range(1, 36):
            ws.column_dimensions[get_column_letter(col)].width = 14

        logger.debug(f"Wrote {df.shape[0]} rows to '{loc_names[loc]}'")

    logger.info(f"Completed writing budget data for {len(loc_names)} locations")


# -----------------------------------------------------------------------------
# Legacy function support (for backward compatibility)
# These functions mimic the old win32com API but use openpyxl internally
# -----------------------------------------------------------------------------

class _ExcelAppProxy:
    """Proxy object that mimics Excel Application for legacy API compatibility."""

    def __init__(self):
        self._workbooks = []
        self.Visible = False
        self.DisplayAlerts = False

    def create_workbook(self):
        """Create a new workbook and track it."""
        wb = create_workbook()
        self._workbooks.append(wb)
        return wb


def excel_init(visible=False, display_alerts=False):
    """Initialize Excel application (legacy API).

    In openpyxl mode, returns a proxy object that can create workbooks.

    Parameters
    ----------
    visible : bool, default=False
        Ignored in openpyxl mode.
    display_alerts : bool, default=False
        Ignored in openpyxl mode.

    Returns
    -------
    excel : _ExcelAppProxy
        Proxy object for creating workbooks.
    """
    logger.debug("excel_init called (openpyxl mode)")
    proxy = _ExcelAppProxy()
    proxy.Visible = visible
    proxy.DisplayAlerts = display_alerts
    return proxy


def excel_new_workbook(excel):
    """Create a new workbook (legacy API).

    Parameters
    ----------
    excel : _ExcelAppProxy or any
        Excel application proxy (ignored, kept for API compatibility).

    Returns
    -------
    workbook : openpyxl.Workbook
        New workbook.
    """
    logger.debug("excel_new_workbook called (openpyxl mode)")
    return create_workbook()


def excel_kill(excel):
    """Close Excel application (legacy API).

    No-op in openpyxl mode - there's no application to close.

    Parameters
    ----------
    excel : any
        Ignored.
    """
    logger.debug("excel_kill called (openpyxl mode - no-op)")
    pass


def xl_quit(excel):
    """Close Excel application (legacy API).

    No-op in openpyxl mode.

    Parameters
    ----------
    excel : any
        Ignored.
    """
    logger.debug("xl_quit called (openpyxl mode - no-op)")
    pass