# test_headall2excel.py
# Unit tests for the headall2excel function in the iwfm package
# Copyright (C) 2026 University of California
# -----------------------------------------------------------------------------
# This information is free; you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This work is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# For a copy of the GNU General Public License, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
# -----------------------------------------------------------------------------

import pytest
import os
import tempfile

# Check if xlsxwriter is available
try:
    import xlsxwriter  # noqa: F401
    del xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

# Check if openpyxl is available for reading Excel files in tests
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import directly from module since it may not be exported in __init__.py
from iwfm.headall2excel import headall2excel, new_excel

# Path to the example C2VSimCG files
EXAMPLE_HEADALL_FILE = os.path.join(
    os.path.dirname(__file__),
    'C2VSimCG-2021',
    'Results',
    'C2VSimCG_GW_HeadAll.out'
)

EXAMPLE_PRE_FILE = os.path.join(
    os.path.dirname(__file__),
    'C2VSimCG-2021',
    'Preprocessor',
    'C2VSimCG_Preprocessor.in'
)

# Check if the example files exist for tests that require them
EXAMPLE_FILES_EXIST = (
    os.path.exists(EXAMPLE_HEADALL_FILE) and
    os.path.exists(EXAMPLE_PRE_FILE)
)


class TestHeadall2ExcelFunctionExists:
    """Test that the headall2excel function exists and is callable."""

    def test_headall2excel_exists(self):
        """Test that headall2excel function exists and is callable."""
        assert headall2excel is not None
        assert callable(headall2excel)

    def test_new_excel_exists(self):
        """Test that new_excel helper function exists and is callable."""
        assert new_excel is not None
        assert callable(new_excel)


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE, reason="xlsxwriter not installed")
class TestNewExcel:
    """Test the new_excel helper function."""

    def test_creates_workbook(self):
        """Test that new_excel creates a workbook object."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')
            workbook = new_excel(outfile)

            assert workbook is not None
            workbook.close()

    def test_creates_file_on_close(self):
        """Test that workbook creates file when closed."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')
            workbook = new_excel(outfile)
            workbook.close()

            assert os.path.exists(outfile)


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE, reason="xlsxwriter not installed")
class TestHeadall2ExcelBasicFunctionality:
    """Test basic functionality of headall2excel."""

    def test_creates_excel_file(self):
        """Test that headall2excel creates an Excel file."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            # Simple test data
            node_coords = [
                [1, 100.0, 200.0],
                [2, 150.0, 250.0],
                [3, 200.0, 300.0],
            ]
            # Data shape: [timesteps][layers][nodes]
            data = [
                [[10.0, 20.0, 30.0], [11.0, 21.0, 31.0]],  # Timestep 1, 2 layers
                [[10.5, 20.5, 30.5], [11.5, 21.5, 31.5]],  # Timestep 2, 2 layers
            ]
            dates = ['09/30/1973', '10/31/1973']
            out_dates = ['09/30/1973', '10/31/1973']

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert os.path.exists(outfile)

    def test_returns_count(self):
        """Test that headall2excel returns the count of worksheets created."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert isinstance(count, int)
            assert count == 1

    def test_returns_correct_count_multiple_dates(self):
        """Test that count matches number of output dates."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]], [[11.0]], [[12.0]]]
            dates = ['09/30/1973', '10/31/1973', '11/30/1973']
            out_dates = ['09/30/1973', '10/31/1973', '11/30/1973']

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 3


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE, reason="xlsxwriter not installed")
class TestHeadall2ExcelDateFiltering:
    """Test date filtering in headall2excel."""

    def test_filters_by_out_dates(self):
        """Test that only specified out_dates are included."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]], [[11.0]], [[12.0]]]
            dates = ['09/30/1973', '10/31/1973', '11/30/1973']
            out_dates = ['10/31/1973']  # Only request one date

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 1

    def test_no_matching_dates_returns_zero(self):
        """Test that no matching dates returns count of zero."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]], [[11.0]]]
            dates = ['09/30/1973', '10/31/1973']
            out_dates = ['12/31/1999']  # Date not in data

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 0

    def test_partial_date_match(self):
        """Test with some dates matching and some not."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]], [[11.0]], [[12.0]]]
            dates = ['09/30/1973', '10/31/1973', '11/30/1973']
            out_dates = ['10/31/1973', '12/31/1999']  # One matches, one doesn't

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 1


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE or not OPENPYXL_AVAILABLE,
                    reason="xlsxwriter or openpyxl not installed")
class TestHeadall2ExcelContent:
    """Test the content of Excel files created by headall2excel."""

    def test_worksheet_has_correct_headers(self):
        """Test that worksheet has correct column headers."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0], [2, 150.0, 250.0]]
            data = [[[10.0, 20.0], [11.0, 21.0]]]  # 2 layers
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            headall2excel(node_coords, data, dates, out_dates, outfile)

            # Read back with openpyxl
            wb = openpyxl.load_workbook(outfile)
            ws = wb.active

            # Check headers
            assert ws.cell(1, 1).value == 'NodeID'
            assert ws.cell(1, 2).value == 'X'
            assert ws.cell(1, 3).value == 'Y'
            assert ws.cell(1, 4).value == 'Layer 1'
            assert ws.cell(1, 5).value == 'Layer 2'

            wb.close()

    def test_worksheet_has_correct_node_data(self):
        """Test that worksheet has correct node coordinate data."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0], [2, 150.0, 250.0]]
            data = [[[10.0, 20.0]]]  # 1 layer
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            headall2excel(node_coords, data, dates, out_dates, outfile)

            wb = openpyxl.load_workbook(outfile)
            ws = wb.active

            # Check first node
            assert ws.cell(2, 1).value == 1  # NodeID
            assert ws.cell(2, 2).value == 100.0  # X
            assert ws.cell(2, 3).value == 200.0  # Y

            # Check second node
            assert ws.cell(3, 1).value == 2
            assert ws.cell(3, 2).value == 150.0
            assert ws.cell(3, 3).value == 250.0

            wb.close()

    def test_worksheet_has_correct_head_data(self):
        """Test that worksheet has correct head values."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0], [2, 150.0, 250.0]]
            data = [[[10.0, 20.0], [11.0, 21.0]]]  # 2 layers, 2 nodes
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            headall2excel(node_coords, data, dates, out_dates, outfile)

            wb = openpyxl.load_workbook(outfile)
            ws = wb.active

            # Check layer 1 values (column 4)
            assert ws.cell(2, 4).value == 10.0  # Node 1, Layer 1
            assert ws.cell(3, 4).value == 20.0  # Node 2, Layer 1

            # Check layer 2 values (column 5)
            assert ws.cell(2, 5).value == 11.0  # Node 1, Layer 2
            assert ws.cell(3, 5).value == 21.0  # Node 2, Layer 2

            wb.close()

    def test_worksheet_names_from_dates(self):
        """Test that worksheet names are derived from dates."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]], [[11.0]]]
            dates = ['09/30/1973', '10/31/1973']
            out_dates = ['09/30/1973', '10/31/1973']

            headall2excel(node_coords, data, dates, out_dates, outfile)

            wb = openpyxl.load_workbook(outfile)

            # Sheet names should have slashes replaced with underscores
            sheet_names = wb.sheetnames
            assert '09_30_1973' in sheet_names
            assert '10_31_1973' in sheet_names

            wb.close()


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE, reason="xlsxwriter not installed")
class TestHeadall2ExcelVerbose:
    """Test the verbose parameter of headall2excel."""

    def test_verbose_false_no_error(self):
        """Test that verbose=False runs without error."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            # Should not raise exception
            headall2excel(node_coords, data, dates, out_dates, outfile, verbose=False)

    def test_verbose_true_no_error(self):
        """Test that verbose=True runs without error."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            # Should not raise exception
            headall2excel(node_coords, data, dates, out_dates, outfile, verbose=True)


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE, reason="xlsxwriter not installed")
class TestHeadall2ExcelEdgeCases:
    """Test edge cases for headall2excel."""

    def test_single_node(self):
        """Test with single node."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 1
            assert os.path.exists(outfile)

    def test_single_layer(self):
        """Test with single layer."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0], [2, 150.0, 250.0]]
            data = [[[10.0, 20.0]]]  # 1 layer, 2 nodes
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 1

    def test_many_layers(self):
        """Test with many layers (4 like C2VSimCG)."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0], [2, 150.0, 250.0]]
            # 4 layers, 2 nodes
            data = [[[10.0, 20.0], [11.0, 21.0], [12.0, 22.0], [13.0, 23.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 1

    def test_empty_out_dates(self):
        """Test with empty out_dates list."""
        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = []  # Empty list

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 0


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE or not EXAMPLE_FILES_EXIST,
                    reason="xlsxwriter not installed or example files not found")
class TestHeadall2ExcelWithRealData:
    """Test headall2excel with real C2VSimCG data."""

    def test_with_c2vsimcg_data(self):
        """Test with actual C2VSimCG data."""
        import iwfm

        with tempfile.TemporaryDirectory() as temp_dir:
            outfile = os.path.join(temp_dir, 'test.xlsx')

            # Read real data
            data, layers, dates, nodes = iwfm.headall_read(EXAMPLE_HEADALL_FILE)

            # Get preprocessor info
            pre_path, _ = os.path.split(EXAMPLE_PRE_FILE)
            pre_files, _ = iwfm.iwfm_read_preproc(EXAMPLE_PRE_FILE)

            # Read node coordinates
            node_file = os.path.join(pre_path, pre_files['node_file'])
            node_coords, node_list, factor = iwfm.iwfm_read_nodes(node_file)

            # Scale coordinates
            for i in range(len(node_coords)):
                node_coords[i][1] *= factor
                node_coords[i][2] *= factor

            # Output first date only
            out_dates = [dates[0]]

            count = headall2excel(node_coords, data, dates, out_dates, outfile)

            assert count == 1
            assert os.path.exists(outfile)
            assert os.path.getsize(outfile) > 0


@pytest.mark.skipif(not XLSXWRITER_AVAILABLE, reason="xlsxwriter not installed")
class TestHeadall2ExcelOutputPath:
    """Test output path handling in headall2excel."""

    def test_output_to_subdirectory(self):
        """Test that output can be written to a subdirectory."""
        with tempfile.TemporaryDirectory() as temp_dir:
            subdir = os.path.join(temp_dir, 'subdir')
            os.makedirs(subdir)
            outfile = os.path.join(subdir, 'test.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            headall2excel(node_coords, data, dates, out_dates, outfile)

            assert os.path.exists(outfile)

    def test_output_with_spaces_in_path(self):
        """Test that output works with spaces in path."""
        with tempfile.TemporaryDirectory() as temp_dir:
            subdir = os.path.join(temp_dir, 'sub dir')
            os.makedirs(subdir)
            outfile = os.path.join(subdir, 'test output.xlsx')

            node_coords = [[1, 100.0, 200.0]]
            data = [[[10.0]]]
            dates = ['09/30/1973']
            out_dates = ['09/30/1973']

            headall2excel(node_coords, data, dates, out_dates, outfile)

            assert os.path.exists(outfile)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
